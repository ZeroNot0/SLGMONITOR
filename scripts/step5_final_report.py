import pandas as pd
from pathlib import Path
import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def run_step5(week_tag: str = None, year: int = None):

    BASE_DIR = Path(__file__).parent.parent
    
    # === 构建输入文件路径 ===
    if week_tag and year:
        INPUT_FILE = BASE_DIR / "intermediate" / str(year) / week_tag / "pivot_table.xlsx"
    else:
        # 向后兼容
        if week_tag and not year:
            pattern = f"*/{week_tag}/pivot_table.xlsx"
            files = list((BASE_DIR / "intermediate").glob(pattern))
            if files:
                INPUT_FILE = sorted(files)[-1]
            else:
                INPUT_FILE = BASE_DIR / "intermediate" / "pivot_table.xlsx"
        else:
            INPUT_FILE = BASE_DIR / "intermediate" / "pivot_table.xlsx"
    
    # === 构建输出文件路径 ===
    if week_tag and year:
        OUTPUT_DIR = BASE_DIR / "output" / str(year)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        OUTPUT_FILE = OUTPUT_DIR / f"{week_tag}_SLG数据监测表.xlsx"
    else:
        OUTPUT_FILE = BASE_DIR / "output" / "SLG数据监测表.xlsx"
    
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    # =====================
    # 读取 STEP4 数据
    # =====================
    df = pd.read_excel(INPUT_FILE)

    # =====================
    # 计算周变动（直接生成字符串列）
    # =====================
    inst_change = (df["当周周安装"] - df["上周周安装"]) / df["上周周安装"]
    rev_change = (df["当周周流水"] - df["上周周流水"]) / df["上周周流水"]

    def arrow_fmt(x):
        if pd.isna(x):
            return ""
        if x >= 0:
            return f"{x*100:.2f}%▲"
        else:
            return f"{x*100:.2f}%▼"

    df["周安装变动"] = inst_change.apply(arrow_fmt)
    df["周流水变动"] = rev_change.apply(arrow_fmt)

    # =====================
    # 删除条件 ①
    # =====================
    cond_delete = (
        (df["当周周安装"] < 400) &
        (df["上周周安装"] < 400) &
        (df["当周周流水"] < 20000) &
        (df["上周周流水"] < 20000)
    )
    df = df[~cond_delete].reset_index(drop=True)

    # =====================
    # 列顺序（含 Unified ID 时保留，供 API 请求使用）
    # =====================
    final_cols = [
        "公司归属",
        "产品归属",
        "第三方记录最早上线时间",
        "当周周安装",
        "上周周安装",
        "周安装变动",
        "当周周流水",
        "上周周流水",
        "周流水变动"
    ]
    if "Unified ID" in df.columns:
        final_cols.insert(final_cols.index("产品归属") + 1, "Unified ID")
    df = df[[c for c in final_cols if c in df.columns]]

    # =====================
    # 输出初始 Excel
    # =====================
    df.to_excel(OUTPUT_FILE, index=False)

    # =====================
    # openpyxl 格式处理
    # =====================
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    # ---- 样式 ----
    blue_header = PatternFill("solid", fgColor="4F81BD")
    blue_summary = PatternFill("solid", fgColor="D9E1F2")
    yellow_row = PatternFill("solid", fgColor="FFF2CC")

    red_font = Font(color="FF0000")
    green_font = Font(color="00B050")
    center_align = Alignment(horizontal="center", vertical="center")

    # ---- 首行深蓝 ----
    for cell in ws[1]:
        cell.fill = blue_header
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = center_align

    # ---- 列索引 ----
    cols = list(df.columns)
    col_company = cols.index("公司归属") + 1
    col_product = cols.index("产品归属") + 1

    col_inst_this = cols.index("当周周安装") + 1
    col_inst_last = cols.index("上周周安装") + 1
    col_inst_chg = cols.index("周安装变动") + 1
    col_rev_this = cols.index("当周周流水") + 1
    col_rev_last = cols.index("上周周流水") + 1
    col_rev_chg = cols.index("周流水变动") + 1

    # ---- 设置列宽 ----
    ws.column_dimensions["B"].width = 38   # 产品列 ≈10cm
    for col in ["A","C","D","E","F","G","H","I"]:
        if col != "B":
            ws.column_dimensions[col].width = 18  # ≈5cm

    # =====================
    # 遍历行进行规则处理
    # =====================
    for r in range(2, ws.max_row + 1):

        company_val = ws.cell(r, col_company).value

        inst_this = ws.cell(r, col_inst_this).value
        inst_last = ws.cell(r, col_inst_last).value
        rev_this = ws.cell(r, col_rev_this).value
        rev_last = ws.cell(r, col_rev_last).value

        # ---- 全行居中 ----
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).alignment = center_align

        # ---- 公司汇总行浅蓝 ----
        if isinstance(company_val, str) and company_val.endswith("汇总"):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = blue_summary
            continue

        # ---- 条件② 小安装高流水 → 产品标红 + 删除线 ----
        if (
            inst_this < 400 and inst_last < 400 and
            (rev_this >= 20000 or rev_last >= 20000)
        ):
            cell = ws.cell(r, col_product)
            cell.font = Font(color="FF0000", strike=True)
            continue  # 已标红的不再参与标黄

        # ---- 条件③ 周安装变动 ≥ +20% → 整行标黄 ----
        val_inst = ws.cell(r, col_inst_chg).value
        if isinstance(val_inst, str) and val_inst.strip() != "":
            if not val_inst.strip().startswith("-"):
                num = float(val_inst.replace("%▲", "").replace("%▼", ""))
                if num >= 20:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(r, c).fill = yellow_row

        # ---- 箭头颜色（字符串首字符判定） ----
        cell_inst = ws.cell(r, col_inst_chg)
        cell_rev = ws.cell(r, col_rev_chg)

        val_inst = cell_inst.value
        val_rev = cell_rev.value

        if isinstance(val_inst, str) and val_inst.strip() != "":
            if val_inst.strip().startswith("-"):
                cell_inst.font = green_font
            else:
                cell_inst.font = red_font

        if isinstance(val_rev, str) and val_rev.strip() != "":
            if val_rev.strip().startswith("-"):
                cell_rev.font = green_font
            else:
                cell_rev.font = red_font

        # ---- 数字格式 ----
        ws.cell(r, col_inst_this).number_format = "#,##0"
        ws.cell(r, col_inst_last).number_format = "#,##0"
        ws.cell(r, col_rev_this).number_format = '"$"#,##0.00'
        ws.cell(r, col_rev_last).number_format = '"$"#,##0.00'

    # ---- 加全边框线 ----
    from openpyxl.styles import Border, Side
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    wb.save(OUTPUT_FILE)

    print("\n🎉 STEP5 完成（最终稳定字符串染色版）")
    print(f"最终输出文件: {OUTPUT_FILE}\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--week", help="例如 0105-0111（可选）")
    parser.add_argument("--year", type=int, help="年份，例如 2025（可选）")
    args = parser.parse_args()

    run_step5(args.week, args.year)


