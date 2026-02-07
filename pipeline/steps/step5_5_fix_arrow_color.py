from pathlib import Path
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Font

def run_step5_5(week_tag: str = None, year: int = None):

    BASE_DIR = Path(__file__).resolve().parent.parent.parent
    
    # === 构建文件路径 ===
    if week_tag and year:
        FILE_PATH = BASE_DIR / "output" / str(year) / f"{week_tag}_SLG数据监测表.xlsx"
    else:
        # 向后兼容
        if week_tag and not year:
            pattern = f"*/{week_tag}_SLG数据监测表.xlsx"
            files = list((BASE_DIR / "output").glob(pattern))
            if files:
                FILE_PATH = sorted(files)[-1]
            else:
                FILE_PATH = BASE_DIR / "output" / "SLG数据监测表.xlsx"
        else:
            FILE_PATH = BASE_DIR / "output" / "SLG数据监测表.xlsx"

    wb = load_workbook(FILE_PATH)
    ws = wb.active

    # ---- 字体颜色 ----
    red_font = Font(color="FF0000")
    green_font = Font(color="00B050")

    # ---- 找到列索引 ----
    headers = [cell.value for cell in ws[1]]

    col_inst_chg = headers.index("周安装变动") + 1
    col_rev_chg  = headers.index("周流水变动") + 1

    # ---- 遍历行重新染色 ----
    for r in range(2, ws.max_row + 1):

        cell_inst = ws.cell(r, col_inst_chg)
        cell_rev  = ws.cell(r, col_rev_chg)

        val_inst = cell_inst.value
        val_rev  = cell_rev.value

        # 周安装变动
        if isinstance(val_inst, str) and val_inst.strip() != "":
            if val_inst.strip().startswith("-"):
                cell_inst.font = green_font
            else:
                cell_inst.font = red_font

        # 周流水变动
        if isinstance(val_rev, str) and val_rev.strip() != "":
            if val_rev.strip().startswith("-"):
                cell_rev.font = green_font
            else:
                cell_rev.font = red_font

    wb.save(FILE_PATH)

    print("\n🎯 STEP5.5 完成：箭头颜色已全部重置")
    print(f"文件已更新: {FILE_PATH}\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--week", help="例如 0105-0111（可选）")
    parser.add_argument("--year", type=int, help="年份，例如 2025（可选）")
    args = parser.parse_args()

    run_step5_5(args.week, args.year)
