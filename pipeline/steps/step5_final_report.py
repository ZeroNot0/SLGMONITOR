import json
import pandas as pd
from pathlib import Path
import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def _default_monitor_rules() -> dict:
    return {
        "version": 1,
        "delete_rules": [
            {
                "conditions": [
                    {"metric": "当周周安装", "op": "<", "value": 400},
                    {"metric": "上周周安装", "op": "<", "value": 400},
                    {"metric": "当周周流水", "op": "<", "value": 20000},
                    {"metric": "上周周流水", "op": "<", "value": 20000},
                ]
            }
        ],
        "strike_rules": [
            {
                "conditions": [
                    {"metric": "当周周安装", "op": "<", "value": 400},
                    {"metric": "上周周安装", "op": "<", "value": 400},
                    {"metric": "当周周流水", "op": ">=", "value": 20000},
                ]
            },
            {
                "conditions": [
                    {"metric": "当周周安装", "op": "<", "value": 400},
                    {"metric": "上周周安装", "op": "<", "value": 400},
                    {"metric": "上周周流水", "op": ">=", "value": 20000},
                ]
            },
        ],
        "yellow_rules": [
            {
                "conditions": [
                    {"metric": "周安装变动", "op": ">=", "value": 20},
                    {"metric": "当周周安装", "op": ">", "value": 1000},
                ]
            }
        ],
        "product_rules": {
            "delete": [],
            "strike": [],
            "yellow": [],
        },
    }


def _load_monitor_rules() -> dict:
    try:
        from app.app_paths import get_data_root
        path = get_data_root() / "config" / "monitor_rules.json"
        if path.is_file():
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
    except Exception:
        pass
    return _default_monitor_rules()


def _normalize_rules(rules: dict) -> dict:
    base = _default_monitor_rules()
    if not isinstance(rules, dict):
        return base
    out = {**base, **rules}
    for key in ("delete_rules", "strike_rules", "yellow_rules"):
        if not isinstance(out.get(key), list):
            out[key] = []
        for idx, rule in enumerate(out[key]):
            if not isinstance(rule, dict):
                continue
            if idx > 0 and rule.get("join") not in ("and", "or"):
                rule["join"] = "or"
    pr = out.get("product_rules") if isinstance(out.get("product_rules"), dict) else {}
    out["product_rules"] = {
        "delete": pr.get("delete") if isinstance(pr.get("delete"), list) else [],
        "strike": pr.get("strike") if isinstance(pr.get("strike"), list) else [],
        "yellow": pr.get("yellow") if isinstance(pr.get("yellow"), list) else [],
    }
    return out


def _norm_str(val) -> str:
    return str(val).strip().lower() if val is not None else ""


def _eval_condition(val, op, target) -> bool:
    try:
        v = float(val)
        t = float(target)
    except Exception:
        return False
    if op == ">":
        return v > t
    if op == ">=":
        return v >= t
    if op == "<":
        return v < t
    if op == "<=":
        return v <= t
    if op == "=":
        return v == t
    if op == "!=":
        return v != t
    return False


def _rule_match(rule: dict, metrics: dict) -> bool:
    conds = rule.get("conditions") if isinstance(rule, dict) else None
    if not conds or not isinstance(conds, list):
        return False
    any_cond = False
    for cond in conds:
        if not isinstance(cond, dict):
            return False
        metric = cond.get("metric")
        op = cond.get("op")
        target = cond.get("value")
        if metric not in metrics:
            return False
        if op not in (">", ">=", "<", "<=", "=", "!="):
            return False
        if target is None or target == "":
            return False
        any_cond = True
        if not _eval_condition(metrics.get(metric), op, target):
            return False
    return any_cond


def _build_product_rule_sets(product_rules: dict):
    def _collect(key, by_value):
        out = set()
        for item in product_rules.get(key, []):
            if not isinstance(item, dict):
                continue
            if item.get("by") != by_value:
                continue
            val = item.get("value")
            if val is None:
                continue
            v = _norm_str(val)
            if v:
                out.add(v)
        return out

    return {
        "delete_name": _collect("delete", "product_name"),
        "delete_id": _collect("delete", "unified_id"),
        "strike_name": _collect("strike", "product_name"),
        "strike_id": _collect("strike", "unified_id"),
        "yellow_name": _collect("yellow", "product_name"),
        "yellow_id": _collect("yellow", "unified_id"),
    }

def run_step5(week_tag: str = None, year: int = None):

    BASE_DIR = Path(__file__).resolve().parent.parent.parent
    
    # === 构建输入文件路径 ===
    if week_tag and year:
        INPUT_FILE = BASE_DIR / "intermediate" / str(year) / week_tag / "pivot_table.xlsx"
    else:
        # 向后兼容
        if week_tag and not year:
            pattern = f"*/{week_tag}/pivot_table.xlsx"
            files = list((BASE_DIR / "intermediate").glob(pattern))
            if files:
                INPUT_FILE = sorted(files)[-1]
            else:
                INPUT_FILE = BASE_DIR / "intermediate" / "pivot_table.xlsx"
        else:
            INPUT_FILE = BASE_DIR / "intermediate" / "pivot_table.xlsx"
    
    # === 构建输出文件路径 ===
    if week_tag and year:
        OUTPUT_DIR = BASE_DIR / "output" / str(year)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        OUTPUT_FILE = OUTPUT_DIR / f"{week_tag}_SLG数据监测表.xlsx"
    else:
        OUTPUT_FILE = BASE_DIR / "output" / "SLG数据监测表.xlsx"
    
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    # =====================
    # 读取 STEP4 数据
    # =====================
    df = pd.read_excel(INPUT_FILE)

    # =====================
    # 计算周变动（直接生成字符串列）
    # =====================
    inst_change = (df["当周周安装"] - df["上周周安装"]) / df["上周周安装"]
    rev_change = (df["当周周流水"] - df["上周周流水"]) / df["上周周流水"]

    def arrow_fmt(x):
        if pd.isna(x):
            return ""
        if x >= 0:
            return f"{x*100:.2f}%▲"
        else:
            return f"{x*100:.2f}%▼"

    df["周安装变动"] = inst_change.apply(arrow_fmt)
    df["周流水变动"] = rev_change.apply(arrow_fmt)

    # =====================
    # 自定义规则：删除 / 划删除线 / 标黄
    # =====================
    rules = _normalize_rules(_load_monitor_rules())
    product_rule_sets = _build_product_rule_sets(rules.get("product_rules", {}))

    df["_周安装变动数值"] = inst_change * 100
    df["_周流水变动数值"] = rev_change * 100

    def _row_metrics(row):
        return {
            "当周周安装": row.get("当周周安装"),
            "上周周安装": row.get("上周周安装"),
            "当周周流水": row.get("当周周流水"),
            "上周周流水": row.get("上周周流水"),
            "周安装变动": row.get("_周安装变动数值"),
            "周流水变动": row.get("_周流水变动数值"),
        }

    def _combine_rule_list(rules: list, metrics: dict) -> bool:
        if not rules:
            return False
        result = _rule_match(rules[0], metrics)
        for rule in rules[1:]:
            join = str((rule or {}).get("join") or "or").strip().lower()
            match = _rule_match(rule, metrics)
            if join == "and":
                result = result and match
            else:
                result = result or match
        return result

    def _match_product_rule(sets_key_name, sets_key_id, product_name, unified_id):
        if product_name and _norm_str(product_name) in product_rule_sets.get(sets_key_name, set()):
            return True
        if unified_id and _norm_str(unified_id) in product_rule_sets.get(sets_key_id, set()):
            return True
        return False

    delete_flags = []
    strike_flags = []
    yellow_flags = []
    for _, row in df.iterrows():
        company_val = row.get("公司归属")
        product_val = row.get("产品归属")
        unified_val = row.get("Unified ID")
        is_summary = isinstance(company_val, str) and company_val.endswith("汇总")
        metrics = _row_metrics(row)
        delete_hit = False
        strike_hit = False
        yellow_hit = False
        if not is_summary:
            delete_hit = _match_product_rule("delete_name", "delete_id", product_val, unified_val) or _combine_rule_list(rules.get("delete_rules", []), metrics)
            strike_hit = _match_product_rule("strike_name", "strike_id", product_val, unified_val) or _combine_rule_list(rules.get("strike_rules", []), metrics)
            yellow_hit = _match_product_rule("yellow_name", "yellow_id", product_val, unified_val) or _combine_rule_list(rules.get("yellow_rules", []), metrics)
        delete_flags.append(bool(delete_hit))
        strike_flags.append(bool(strike_hit))
        yellow_flags.append(bool(yellow_hit))

    if delete_flags:
        df = df[~pd.Series(delete_flags, index=df.index)].reset_index(drop=True)
        strike_flags = [v for v, keep in zip(strike_flags, delete_flags) if not keep]
        yellow_flags = [v for v, keep in zip(yellow_flags, delete_flags) if not keep]

    # =====================
    # 列顺序：保留 Unified ID 供 target / build_final_join 匹配地区数据，其余为产品归属与指标列
    # =====================
    final_cols = [
        "公司归属",
        "产品归属",
        "Unified ID",
        "第三方记录最早上线时间",
        "当周周安装",
        "上周周安装",
        "周安装变动",
        "当周周流水",
        "上周周流水",
        "周流水变动"
    ]
    df = df[[c for c in final_cols if c in df.columns]]

    # =====================
    # 输出初始 Excel
    # =====================
    df.to_excel(OUTPUT_FILE, index=False)

    # =====================
    # openpyxl 格式处理
    # =====================
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    # ---- 样式 ----
    blue_header = PatternFill("solid", fgColor="4F81BD")
    blue_summary = PatternFill("solid", fgColor="D9E1F2")
    yellow_row = PatternFill("solid", fgColor="FFF2CC")

    red_font = Font(color="FF0000")
    green_font = Font(color="00B050")
    center_align = Alignment(horizontal="center", vertical="center")

    # ---- 首行深蓝 ----
    for cell in ws[1]:
        cell.fill = blue_header
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = center_align

    # ---- 列索引 ----
    cols = list(df.columns)
    col_company = cols.index("公司归属") + 1
    col_product = cols.index("产品归属") + 1

    col_inst_this = cols.index("当周周安装") + 1
    col_inst_last = cols.index("上周周安装") + 1
    col_inst_chg = cols.index("周安装变动") + 1
    col_rev_this = cols.index("当周周流水") + 1
    col_rev_last = cols.index("上周周流水") + 1
    col_rev_chg = cols.index("周流水变动") + 1

    # ---- 设置列宽 ----
    ws.column_dimensions["B"].width = 38   # 产品列 ≈10cm
    for col in ["A", "C", "D", "E", "F", "G", "H", "I", "J"]:
        if col != "B":
            ws.column_dimensions[col].width = 18  # ≈5cm

    # =====================
    # 遍历行进行规则处理
    # =====================
    for r in range(2, ws.max_row + 1):

        company_val = ws.cell(r, col_company).value

        inst_this = ws.cell(r, col_inst_this).value
        inst_last = ws.cell(r, col_inst_last).value
        rev_this = ws.cell(r, col_rev_this).value
        rev_last = ws.cell(r, col_rev_last).value

        # ---- 全行居中 ----
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).alignment = center_align

        # ---- 公司汇总行浅蓝 ----
        if isinstance(company_val, str) and company_val.endswith("汇总"):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = blue_summary
            continue

        # ---- 自定义规则：划删除线 / 标黄 ----
        idx = r - 2
        if idx < len(strike_flags) and strike_flags[idx]:
            cell = ws.cell(r, col_product)
            cell.font = Font(color="FF0000", strike=True)

        if idx < len(yellow_flags) and yellow_flags[idx]:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = yellow_row

        # ---- 箭头颜色（字符串首字符判定） ----
        cell_inst = ws.cell(r, col_inst_chg)
        cell_rev = ws.cell(r, col_rev_chg)

        val_inst = cell_inst.value
        val_rev = cell_rev.value

        if isinstance(val_inst, str) and val_inst.strip() != "":
            if val_inst.strip().startswith("-"):
                cell_inst.font = green_font
            else:
                cell_inst.font = red_font

        if isinstance(val_rev, str) and val_rev.strip() != "":
            if val_rev.strip().startswith("-"):
                cell_rev.font = green_font
            else:
                cell_rev.font = red_font

        # ---- 数字格式 ----
        ws.cell(r, col_inst_this).number_format = "#,##0"
        ws.cell(r, col_inst_last).number_format = "#,##0"
        ws.cell(r, col_rev_this).number_format = '"$"#,##0.00'
        ws.cell(r, col_rev_last).number_format = '"$"#,##0.00'

    # ---- 加全边框线 ----
    from openpyxl.styles import Border, Side
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                             min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    wb.save(OUTPUT_FILE)

    print("\n🎉 STEP5 完成（最终稳定字符串染色版）")
    print(f"最终输出文件: {OUTPUT_FILE}\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--week", help="例如 0105-0111（可选）")
    parser.add_argument("--year", type=int, help="年份，例如 2025（可选）")
    args = parser.parse_args()

    run_step5(args.week, args.year)


