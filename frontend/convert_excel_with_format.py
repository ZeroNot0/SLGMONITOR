#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将Excel文件转换为JSON，保留格式信息（颜色、字体等）
"""

import json
import os
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

BASE_DIR = Path(__file__).parent.parent


def _get_data_dir() -> Path:
    override = os.environ.get("SLG_MONITOR_DATA_DIR", "").strip()
    if override:
        return Path(override).expanduser().resolve() / "frontend" / "data"
    appdata = os.environ.get("APPDATA")
    if appdata:
        return Path(appdata).expanduser().resolve() / "SLGMonitor" / "frontend" / "data"
    try:
        from app.app_paths import get_data_root
        return get_data_root() / "frontend" / "data"
    except Exception:
        return Path(__file__).resolve().parent / "data"


DATA_DIR = _get_data_dir()
TARGET_ROW_BG = "#FFF2CC"  # 标黄行黄底（由规则决定）
SUMMARY_ROW_BG = "#D9E1F2"  # 汇总行浅蓝


def _default_monitor_rules() -> dict:
    return {
        "version": 1,
        "delete_rules": [
            {
                "conditions": [
                    {"metric": "当周周安装", "op": "<", "value": 400},
                    {"metric": "上周周安装", "op": "<", "value": 400},
                    {"metric": "当周周流水", "op": "<", "value": 20000},
                    {"metric": "上周周流水", "op": "<", "value": 20000},
                ]
            }
        ],
        "strike_rules": [
            {
                "conditions": [
                    {"metric": "当周周安装", "op": "<", "value": 400},
                    {"metric": "上周周安装", "op": "<", "value": 400},
                    {"metric": "当周周流水", "op": ">=", "value": 20000},
                ]
            },
            {
                "conditions": [
                    {"metric": "当周周安装", "op": "<", "value": 400},
                    {"metric": "上周周安装", "op": "<", "value": 400},
                    {"metric": "上周周流水", "op": ">=", "value": 20000},
                ]
            },
        ],
        "yellow_rules": [
            {
                "conditions": [
                    {"metric": "周安装变动", "op": ">=", "value": 20},
                    {"metric": "当周周安装", "op": ">", "value": 1000},
                ]
            }
        ],
        "product_rules": {
            "delete": [],
            "strike": [],
            "yellow": [],
        },
    }


def _load_monitor_rules() -> dict:
    try:
        from app.app_paths import get_data_root
        path = get_data_root() / "config" / "monitor_rules.json"
        if path.is_file():
            data = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return data
    except Exception:
        pass
    return _default_monitor_rules()


def _normalize_rules(rules: dict) -> dict:
    base = _default_monitor_rules()
    if not isinstance(rules, dict):
        return base
    out = {**base, **rules}
    for key in ("delete_rules", "strike_rules", "yellow_rules"):
        if not isinstance(out.get(key), list):
            out[key] = []
        for idx, rule in enumerate(out[key]):
            if not isinstance(rule, dict):
                continue
            if idx > 0 and rule.get("join") not in ("and", "or"):
                rule["join"] = "or"
    pr = out.get("product_rules") if isinstance(out.get("product_rules"), dict) else {}
    out["product_rules"] = {
        "delete": pr.get("delete") if isinstance(pr.get("delete"), list) else [],
        "strike": pr.get("strike") if isinstance(pr.get("strike"), list) else [],
        "yellow": pr.get("yellow") if isinstance(pr.get("yellow"), list) else [],
    }
    return out


def _norm_str(val) -> str:
    return str(val).strip().lower() if val is not None else ""


def _eval_condition(val, op, target) -> bool:
    try:
        v = float(val)
        t = float(target)
    except Exception:
        return False
    if op == ">":
        return v > t
    if op == ">=":
        return v >= t
    if op == "<":
        return v < t
    if op == "<=":
        return v <= t
    if op == "=":
        return v == t
    if op == "!=":
        return v != t
    return False


def _rule_match(rule: dict, metrics: dict) -> bool:
    conds = rule.get("conditions") if isinstance(rule, dict) else None
    if not conds or not isinstance(conds, list):
        return False
    any_cond = False
    for cond in conds:
        if not isinstance(cond, dict):
            return False
        metric = cond.get("metric")
        op = cond.get("op")
        target = cond.get("value")
        if metric not in metrics:
            return False
        if op not in (">", ">=", "<", "<=", "=", "!="):
            return False
        if target is None or target == "":
            return False
        any_cond = True
        if not _eval_condition(metrics.get(metric), op, target):
            return False
    return any_cond


def _build_product_rule_sets(product_rules: dict):
    def _collect(key, by_value):
        out = set()
        for item in product_rules.get(key, []):
            if not isinstance(item, dict):
                continue
            if item.get("by") != by_value:
                continue
            val = item.get("value")
            if val is None:
                continue
            v = _norm_str(val)
            if v:
                out.add(v)
        return out

    return {
        "yellow_name": _collect("yellow", "product_name"),
        "yellow_id": _collect("yellow", "unified_id"),
    }


def _parse_install_change(val):
    """从「周安装变动」字符串（如 31.81%▲、-16.51%▼）解析出数值，无法解析返回 None。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if not s:
        return None
    import re
    m = re.search(r"([+-]?\d+\.?\d*)\s*%", s)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def _parse_install_count(val):
    """当周周安装：转为数字，无法解析返回 0。"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0
    try:
        return float(str(val).strip().replace("$", "").replace(",", ""))
    except (ValueError, TypeError):
        return 0


def get_cell_style(cell):
    """获取单元格的样式信息"""
    style = {
        'font_color': None,
        'bg_color': None,
        'bold': False,
        'text': str(cell.value) if cell.value is not None else ''
    }
    
    # 字体颜色
    try:
        if cell.font and cell.font.color:
            rgb = cell.font.color.rgb
            if rgb:
                # 转换为字符串，处理不同格式
                if isinstance(rgb, str):
                    style['font_color'] = rgb
                else:
                    # 可能是RGB对象，尝试获取值
                    style['font_color'] = str(rgb)
    except:
        pass
    
    # 背景颜色（只保留非默认颜色）
    try:
        if cell.fill and cell.fill.start_color:
            rgb = cell.fill.start_color.rgb
            if rgb:
                rgb_str = str(rgb) if isinstance(rgb, str) else str(rgb)
                # 过滤掉默认的白色和黑色背景
                # 00000000 = 透明/默认白色, FFFFFFFF = 白色, 000000 = 黑色
                if rgb_str and rgb_str not in ['00000000', 'FFFFFFFF', '00FFFFFF', '000000', 'FF000000', '00FFFFFF']:
                    # 检查是否是接近黑色的颜色（可能是默认值）
                    if rgb_str.startswith('00') and len(rgb_str) == 8:
                        # 检查RGB值，如果是纯黑或接近纯黑，跳过
                        hex_part = rgb_str[2:] if rgb_str.startswith('00') else rgb_str
                        if hex_part and hex_part != '000000' and hex_part != 'FFFFFF':
                            style['bg_color'] = rgb_str
                    elif not rgb_str.startswith('00') or (rgb_str.startswith('00') and rgb_str[2:] not in ['000000', 'FFFFFF']):
                        style['bg_color'] = rgb_str
    except:
        pass
    
    # 粗体
    try:
        if cell.font:
            style['bold'] = bool(cell.font.bold)
    except:
        pass
    
    return style


def convert_excel_to_json_with_format(year, week_tag):
    """将Excel文件转换为JSON，保留格式信息"""
    excel_file = BASE_DIR / "output" / str(year) / f"{week_tag}_SLG数据监测表.xlsx"
    json_file = DATA_DIR / str(year) / f"{week_tag}_formatted.json"
    
    if not excel_file.exists():
        raise FileNotFoundError(f"Excel文件不存在: {excel_file}")
    
    # 使用openpyxl读取格式
    wb = load_workbook(excel_file, data_only=False)
    ws = wb.active
    
    # 读取数据和格式
    data = {
        "headers": [],
        "rows": [],
        "styles": []  # 存储每个单元格的样式
    }
    
    # 读取表头
    header_row = []
    header_styles = []
    for col_idx, cell in enumerate(ws[1], 1):
        header_row.append(str(cell.value) if cell.value is not None else '')
        header_styles.append(get_cell_style(cell))
    data["headers"] = header_row
    data["styles"].append(header_styles)
    
    # 读取数据行
    for row_idx in range(2, ws.max_row + 1):
        row_data = []
        row_styles = []
        for col_idx, cell in enumerate(ws[row_idx], 1):
            row_data.append(str(cell.value) if cell.value is not None else '')
            row_styles.append(get_cell_style(cell))
        data["rows"].append(row_data)
        data["styles"].append(row_styles)
    
    # 标黄规则：按用户配置；汇总行保持浅蓝
    rules = _normalize_rules(_load_monitor_rules())
    product_rule_sets = _build_product_rule_sets(rules.get("product_rules", {}))
    headers = data["headers"]
    col_company = headers.index("公司归属") if "公司归属" in headers else -1
    col_product = headers.index("产品归属") if "产品归属" in headers else -1
    col_uid = headers.index("Unified ID") if "Unified ID" in headers else -1
    col_inst_this = headers.index("当周周安装") if "当周周安装" in headers else -1
    col_inst_last = headers.index("上周周安装") if "上周周安装" in headers else -1
    col_rev_this = headers.index("当周周流水") if "当周周流水" in headers else -1
    col_rev_last = headers.index("上周周流水") if "上周周流水" in headers else -1
    col_inst_chg = headers.index("周安装变动") if "周安装变动" in headers else -1
    col_rev_chg = headers.index("周流水变动") if "周流水变动" in headers else -1

    def _combine_rule_list(rule_list, metrics):
        if not rule_list:
            return False
        result = _rule_match(rule_list[0], metrics)
        for rule in rule_list[1:]:
            join = str((rule or {}).get("join") or "or").strip().lower()
            match = _rule_match(rule, metrics)
            if join == "and":
                result = result and match
            else:
                result = result or match
        return result

    def _match_product_rule(product_name, unified_id):
        if product_name and _norm_str(product_name) in product_rule_sets.get("yellow_name", set()):
            return True
        if unified_id and _norm_str(unified_id) in product_rule_sets.get("yellow_id", set()):
            return True
        return False

    for row_idx, row in enumerate(data["rows"]):
        style_row = data["styles"][row_idx + 1]  # styles[0] 为表头
        company_val = str(row[col_company]).strip() if col_company >= 0 and col_company < len(row) else ""
        is_summary = company_val.endswith("汇总")
        if is_summary:
            for s in style_row:
                s["bg_color"] = SUMMARY_ROW_BG
            continue
        product_val = str(row[col_product]).strip() if col_product >= 0 and col_product < len(row) else ""
        uid_val = str(row[col_uid]).strip() if col_uid >= 0 and col_uid < len(row) else ""
        inst_this = _parse_install_count(row[col_inst_this]) if col_inst_this >= 0 and col_inst_this < len(row) else 0
        inst_last = _parse_install_count(row[col_inst_last]) if col_inst_last >= 0 and col_inst_last < len(row) else 0
        rev_this = _parse_install_count(row[col_rev_this]) if col_rev_this >= 0 and col_rev_this < len(row) else 0
        rev_last = _parse_install_count(row[col_rev_last]) if col_rev_last >= 0 and col_rev_last < len(row) else 0
        inst_chg = _parse_install_change(row[col_inst_chg]) if col_inst_chg >= 0 and col_inst_chg < len(row) else None
        rev_chg = _parse_install_change(row[col_rev_chg]) if col_rev_chg >= 0 and col_rev_chg < len(row) else None
        metrics = {
            "当周周安装": inst_this,
            "上周周安装": inst_last,
            "当周周流水": rev_this,
            "上周周流水": rev_last,
            "周安装变动": inst_chg,
            "周流水变动": rev_chg,
        }
        yellow_hit = _match_product_rule(product_val, uid_val) or _combine_rule_list(rules.get("yellow_rules", []), metrics)
        if yellow_hit:
            for s in style_row:
                if not s.get("bg_color"):
                    s["bg_color"] = TARGET_ROW_BG
    
    # 保存JSON
    json_file.parent.mkdir(parents=True, exist_ok=True)
    
    # 将RGB颜色转换为可序列化的格式
    def serialize_style(style):
        result = {}
        if style.get('font_color'):
            font_color = str(style['font_color'])
            # 处理openpyxl的RGB格式：00FFFFFF -> #FFFFFF
            if len(font_color) == 8 and font_color.startswith('00'):
                font_color = '#' + font_color[2:]
            elif len(font_color) == 6:
                font_color = '#' + font_color
            elif not font_color.startswith('#'):
                font_color = '#' + font_color
            # 只保留非默认的字体颜色（排除黑色和白色）
            if font_color not in ['#000000', '#FFFFFF', '#00000000', '#FFFFFFFF']:
                result['font_color'] = font_color
        
        if style.get('bg_color'):
            bg_color = str(style['bg_color'])
            # 处理openpyxl的RGB格式
            if len(bg_color) == 8 and bg_color.startswith('00'):
                bg_color = '#' + bg_color[2:]
            elif len(bg_color) == 6:
                bg_color = '#' + bg_color
            elif not bg_color.startswith('#'):
                bg_color = '#' + bg_color
            # 只保留有意义的背景色（排除黑色、白色、透明）
            if bg_color not in ['#000000', '#FFFFFF', '#00000000', '#FFFFFFFF', '#00FFFFFF', '#FF000000']:
                result['bg_color'] = bg_color
        
        if style.get('bold'):
            result['bold'] = True
        
        return result if result else None
    
    # 序列化样式
    serialized_styles = []
    for row_styles in data["styles"]:
        serialized_styles.append([serialize_style(s) for s in row_styles])
    data["styles"] = serialized_styles
    
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"✅ 转换完成: {json_file}")
    print(f"   数据行数: {len(data['rows'])}")
    print(f"   列数: {len(data['headers'])}")
    return json_file


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, required=True, help="年份")
    parser.add_argument("--week", required=True, help="周标签，例如 1201-1207")
    args = parser.parse_args()
    
    convert_excel_to_json_with_format(args.year, args.week)
