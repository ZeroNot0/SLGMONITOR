#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将 intermediate/{年}/{周}/metrics_total.xlsx 转为 frontend/data/{年}/{周}/metrics_total.json，
供产品详情页读取「All Time Downloads (WW)」「All Time Revenue (WW)」等关键数据。
与 _formatted.json 同结构：{ "headers": [...], "rows": [[...], ...] }。
使用 openpyxl 读取，与 convert_excel_with_format 一致。

运行示例（使用 Anaconda 虚拟环境 deeppython）:
  conda activate deeppython
  python frontend/convert_metrics_to_json.py --year 2026 --week 0112-0118
"""
import json
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent.parent
INTERMEDIATE_DIR = BASE_DIR / "intermediate"
DATA_DIR = Path(__file__).resolve().parent / "data"


def _cell_value(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return v
    return str(v).strip() if v else ""


def convert_metrics_to_json(year: int, week_tag: str) -> bool:
    xlsx_path = INTERMEDIATE_DIR / str(year) / week_tag / "metrics_total.xlsx"
    if not xlsx_path.exists() or xlsx_path.stat().st_size == 0:
        return False
    out_dir = DATA_DIR / str(year) / week_tag
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = out_dir / "metrics_total.json"

    try:
        # read_only + iter_rows 流式读取，7 万多行时比逐格 cell() 快很多
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ⚠️ 无法读取 {xlsx_path.name}（{e}），跳过")
        return False
    try:
        ws = wb.active
        row_iter = ws.iter_rows(values_only=True)
        first = next(row_iter, None)
        if not first:
            return False
        headers = [_cell_value(v) for v in first]
        rows = []
        for row in row_iter:
            row_list = []
            for v in row:
                if v is None:
                    row_list.append("")
                elif isinstance(v, (int, float)):
                    row_list.append(v)
                else:
                    row_list.append(str(v).strip() if v else "")
            rows.append(row_list)
    finally:
        wb.close()

    data = {"headers": headers, "rows": rows}
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  ✅ 已生成: {json_path.relative_to(BASE_DIR)}（{len(rows)} 行）")
    return True


def main():
    import argparse
    parser = argparse.ArgumentParser(description="将 metrics_total.xlsx 转为 metrics_total.json")
    parser.add_argument("--year", type=int, required=True, help="年份，如 2026")
    parser.add_argument("--week", type=str, required=True, help="周标签，如 0119-0125")
    args = parser.parse_args()
    if not convert_metrics_to_json(args.year, args.week):
        print(f"  ⏭ 未找到 intermediate/{args.year}/{args.week}/metrics_total.xlsx，跳过")
        raise SystemExit(0)


if __name__ == "__main__":
    main()
